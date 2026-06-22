# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, RadarChart
from openpyxl.chart.label import DataLabelList

FONT="Arial"; NAVY="1F3864"; POOR="C00000"; RICH="2E5496"
HDR=PatternFill("solid",fgColor="1F3864")
POORH=PatternFill("solid",fgColor="F4CCCC")
RICHH=PatternFill("solid",fgColor="D9E1F2")
CAT=PatternFill("solid",fgColor="EDEDED")
TOT=PatternFill("solid",fgColor="1F3864")
AVG=PatternFill("solid",fgColor="FFF2CC")
# scale box gradient poor->rich (B..F)
BOX=[PatternFill("solid",fgColor=c) for c in ["F4CCCC","FBE0E0","F2F2F2","E4EAF6","D9E1F2"]]
thin=Side(style="thin",color="9BA7B5"); border=Border(thin,thin,thin,thin)
med=Side(style="medium",color="7F7F7F"); box_border=Border(med,med,med,med)

cats_data=[
 ("1. תפיסה עצמית ומקור הפרנסה",[
   ('"אני איש/אשת מקצוע" – הכסף שלי מגיע מהמקצוע שלי.','"אני בעל/ת עסק" – הכסף שלי מגיע מהעסק ומהנכסים שלי.'),
   ("חשוב לי ללמוד איך כותבים קורות חיים ומתכוננים לראיון עבודה.","חשוב לי ללמוד איך כותבים תוכנית עסקית, מגייסים משקיעים ורוכשים חברות."),
   ("השאיפה שלי היא להתקדם בסולם הדרגות.","השאיפה שלי היא לשלוט בסולם / לבנות משהו משלי."),
 ]),
 ("2. השכלה והשקעה עצמית",[
   ("משקיע/ה בהשכלה המקצועית שלי, אבל פחות בעצמי ובידע על כסף.","משקיע/ה בעצמי, בחינוך הפיננסי ובמיומנויות של כסף וניהול."),
 ]),
 ("3. זמן מול כסף",[
   ('"הזמן שלי שווה כסף" – אני מחליף/ה זמן בכסף (מוכר/ת שעות).','"הזמן שלי יקר מכסף" – אני מחליף/ה זמן ביצירת נכסים והכנסה פאסיבית.'),
 ]),
 ("4. צמיחה ולמידה",[
   ("בתחומים שאני מכיר/ה אני כבר יודע/ת מספיק, אין הרבה מה ללמוד.",'תמיד יש לי מה ללמוד – "His cup is never full".'),
   ("כשמשהו לא מצליח, כנראה שזה פשוט לא בשבילי.","כשמשהו לא מצליח, אני בודק/ת מה לשנות בעצמי."),
 ]),
 ("5. תודעת שפע מול מחסור",[
   ("כמות הכסף בעולם מוגבלת – אם אחד מרוויח, אחר מפסיד.","יש שפע והזדמנויות לכולם – הכסף גדל ככל שיש יותר שחקנים."),
   ('"אני לא יכול/ה להרשות לעצמי את זה".','"איך אוכל להרשות לעצמי את זה?".'),
   ("כשמגיעה הזדמנות אני חשדן/ית ונרתע/ת.","כשמגיעה הזדמנות אני פועל/ת ומנצל/ת אותה."),
 ]),
 ("6. תפקיד הכסף – שמירה מול יצירה",[
   ("כסף נועד בעיקר להישמר ולהיחסך.","כסף נועד לזרום, להיות מושקע ולבנות."),
 ]),
 ("7. אסטרטגיה: הגנה מול התקפה",[
   ('משחק/ת "הגנה" – לא רוצה להפסיד, נמנע/ת מסיכונים, משחק/ת בטוח.','משחק/ת "התקפה" – רוצה לנצח, לוקח/ת סיכונים מחושבים.'),
   ("פועל/ת לפי מה שכולם עושים ומתייעץ/ת עם הסביבה.",'פועל/ת מתוך ידע ולמידה עצמאית (בד"כ הפוך מהרוב).'),
 ]),
 ("8. אחריות ושליטה",[
   ("ההצלחה הכלכלית שלי תלויה בגורמים שאין לי שליטה עליהם.","ההצלחה הכלכלית שלי תלויה בעיקר בי."),
   ("כשמשהו משתבש אני מחפש/ת את הסיבה החיצונית / האשם.","כשמשהו משתבש אני לוקח/ת אחריות ומחפש/ת מה לשנות."),
 ]),
 ("9. תגמול לפי ערך",[
   ("מעדיף/ה שכר קבוע לפי זמן ושעות.","מעדיף/ה תגמול לפי הערך והתוצאות שאני מביא/ה."),
 ]),
 ("10. ערך עצמי וביטחון",[
   ("לא תמיד בטוח/ה ביכולות שלי ובתוצאות שאוכל להשיג.","בטוח/ה ביכולות שלי ובתוצאות שאני מסוגל/ת להשיג."),
   ("הביטחון שלי נשען על המקצוע / התואר שלי.","הביטחון שלי נשען על היכולות שלי."),
 ]),
 ("11. מיקוד ביעד מול פחד מכישלון",[
   ("פחד מכישלון עוצר אותי – אפעל רק כשהכול מושלם.","מתמקד/ת ביעד, יוצא/ת לדרך גם כשלא הכול מושלם."),
   ("ניתוח יתר נוטה לשתק אותי (analysis paralysis).","הניתוח עוזר לי לבחור, להבין סיכונים ולפעול."),
 ]),
 ("12. חשיבה בגדול (כמויות)",[
   ('חושב/ת ועובד/ת בקטן, בעיקר "אחד על אחד".',"חושב/ת בגדול, מחפש/ת להגדיל את הכמות וההשפעה."),
 ]),
 ("13. סביבה ורגשות כלפי עשירים",[
   ("כלפי אנשים עשירים אני נוטה להרגיש כעס / קנאה / ביקורת.","כלפי אנשים עשירים אני מרגיש/ה הערצה והשראה."),
   ("נוח לי בסביבה של אנשים עם מיינדסט דומה לשלי.","שואף/ת להקיף את עצמי באנשים שהגיעו לאן שאני רוצה."),
 ]),
]

wb=Workbook()
def w(ws,cell,val,size=11,bold=False,color="000000",wrap=True,fill=None,align="right",valign="center"):
    c=ws[cell]; c.value=val
    c.font=Font(name=FONT,size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal=align,vertical=valign,wrap_text=wrap)
    if fill:c.fill=fill
    return c

# ===== הוראות =====
ws0=wb.active; ws0.title="הוראות"; ws0.sheet_view.rightToLeft=True
ws0.column_dimensions['A'].width=3; ws0.column_dimensions['B'].width=108
w(ws0,"B2","שאלון מיפוי עצמי: מיינדסט עני מול מיינדסט עשיר",18,True,NAVY)
w(ws0,"B3",'"The other side of the coin"',12,False,RICH)
w(ws0,"B5","איך ממלאים?",13,True,NAVY)
w(ws0,"B6","בלשונית \"שאלון\", כל שורה מציגה שני ניסוחים מנוגדים: מימין ניסוח של מיינדסט עני, משמאל ניסוח של מיינדסט עשיר.")
w(ws0,"B7","ביניהם יש סקאלה של חמש משבצות. סמנו משבצת אחת (בחרו ✗ מהרשימה הנפתחת) לפי המיקום שמתאר אתכם:")
w(ws0,"B8","המשבצת הימנית = מתחבר/ת לחלוטין לניסוח העני · האמצעית = שילוב · השמאלית = מתחבר/ת לחלוטין לניסוח העשיר.",11,True)
w(ws0,"B9","אין מספרים על הסקאלה בכוונה – פשוט סמנו לאן אתם נוטים. שני הצדדים מוצגים שווה, אין תשובה \"נכונה\".")
w(ws0,"B11","איפה רואים את התוצאה?",13,True,NAVY)
w(ws0,"B12","בלשונית \"תוצאות\" יחושבו אוטומטית ממוצע לכל קטגוריה, ציון כולל, ושני גרפים. בלשונית השאלון עצמה אין ציונים – רק הסימון.")
w(ws0,"B13","1.0–2.0  →  מיינדסט עני דומיננטי   |   2.1–3.0  →  נטייה לעני   |   3.1–4.0  →  נטייה לעשיר   |   4.1–5.0  →  עשיר דומיננטי")
w(ws0,"B15","טיפ: התמקדו בקטגוריות עם הציון הנמוך ביותר – שם הפוטנציאל הגדול ביותר לצמיחה.",11,True,POOR)

# ===== שאלון =====
ws=wb.create_sheet("שאלון"); ws.sheet_view.rightToLeft=True
ws.column_dimensions['A'].width=50
for col in "BCDEF": ws.column_dimensions[col].width=6
ws.column_dimensions['G'].width=50
w(ws,"A1","שאלון מיפוי עצמי – מיינדסט עני מול מיינדסט עשיר",15,True,"FFFFFF",fill=HDR,align="center")
ws.merge_cells("A1:G1"); ws.row_dimensions[1].height=26
for col in "ABCDEFG": ws[f"{col}1"].fill=HDR
w(ws,"A2","סמנו ✗ במשבצת אחת בכל שורה. ימין = מתחבר/ת לעני  ·  אמצע = שילוב  ·  שמאל = מתחבר/ת לעשיר",11,True,NAVY,align="center")
ws.merge_cells("A2:G2"); ws.row_dimensions[2].height=20
hdr=3
w(ws,f"A{hdr}","מיינדסט עני",12,True,POOR,fill=POORH,align="center")
labels=["עני\nמאוד","עני","באמצע","עשיר","עשיר\nמאוד"]
for i,col in enumerate("BCDEF"):
    w(ws,f"{col}{hdr}",labels[i],8,True,"595959",fill=BOX[i],align="center")
    ws[f"{col}{hdr}"].border=border
w(ws,f"G{hdr}","מיינדסט עשיר",12,True,RICH,fill=RICHH,align="center")
ws["A3"].border=border; ws["G3"].border=border
ws.row_dimensions[hdr].height=28

dv=DataValidation(type="list",formula1='"✗"',allow_blank=True,showErrorMessage=True,
                  errorTitle="סימון",error="בחרו ✗ מהרשימה (או השאירו ריק)")
dv.prompt="בחרו ✗ לסימון המשבצת"; dv.promptTitle="סימון"; ws.add_data_validation(dv)

r=hdr+1
item_rows=[]      # (category_index, sheet_row)
cat_ranges=[]     # (cat_name, [sheet_rows])
for ci,(cat_name,pairs) in enumerate(cats_data):
    w(ws,f"A{r}",cat_name,12,True,NAVY,fill=CAT,align="center")
    ws.merge_cells(f"A{r}:G{r}")
    for col in "ABCDEFG": ws[f"{col}{r}"].fill=CAT; ws[f"{col}{r}"].border=border
    ws.row_dimensions[r].height=20; r+=1
    rows_here=[]
    for poor,rich in pairs:
        w(ws,f"A{r}",poor,11,color="000000")
        for i,col in enumerate("BCDEF"):
            c=ws[f"{col}{r}"]; c.fill=BOX[i]; c.border=box_border
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.font=Font(name=FONT,size=12,bold=True,color=POOR if i<2 else (RICH if i>2 else "595959"))
            dv.add(c)
        w(ws,f"G{r}",rich,11,color="000000")
        ws[f"A{r}"].border=border; ws[f"G{r}"].border=border
        ws.row_dimensions[r].height=44
        rows_here.append(r); item_rows.append(r); r+=1
    cat_ranges.append((cat_name,rows_here))
ws.freeze_panes="A4"

# ===== חישוב (hidden) =====
wsc=wb.create_sheet("חישוב")
wsc["A1"]="שורת שאלון"; wsc["B1"]="ציון (1-5)"
cat_calc=[]  # (cat_name, calc_start, calc_end)
cr=2
for cat_name,rows_here in cat_ranges:
    s=cr
    for sr in rows_here:
        wsc[f"A{cr}"]=sr
        wsc[f"B{cr}"]=(f'=IF(COUNTA(שאלון!B{sr}:F{sr})=0,"",'
                       f'1*(שאלון!B{sr}<>"")+2*(שאלון!C{sr}<>"")+3*(שאלון!D{sr}<>"")'
                       f'+4*(שאלון!E{sr}<>"")+5*(שאלון!F{sr}<>""))')
        cr+=1
    cat_calc.append((cat_name,s,cr-1))
calc_last=cr-1
wsc.sheet_state="hidden"

# ===== תוצאות =====
ws2=wb.create_sheet("תוצאות"); ws2.sheet_view.rightToLeft=True
ws2.column_dimensions['A'].width=42; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=22
w(ws2,"A1","סיכום תוצאות לפי קטגוריה",15,True,"FFFFFF",fill=HDR,align="center")
ws2.merge_cells("A1:C1"); ws2.row_dimensions[1].height=24
for col in "ABC": ws2[f"{col}1"].fill=HDR
w(ws2,"A2","קטגוריה",11,True,"FFFFFF",fill=HDR)
w(ws2,"B2","ממוצע",11,True,"FFFFFF",fill=HDR,align="center")
w(ws2,"C2","פירוש",11,True,"FFFFFF",fill=HDR,align="center")
for col in "ABC": ws2[f"{col}2"].border=border
def interp(cellref,rr):
    return (f'=IF({cellref}="","-",IF({cellref}<=2,"עני דומיננטי",IF({cellref}<=3,"נטייה לעני",'
            f'IF({cellref}<=4,"נטייה לעשיר","עשיר דומיננטי"))))')
rr=3; first_cat_row=3
for cat_name,s,e in cat_calc:
    w(ws2,f"A{rr}",cat_name,11)
    bc=ws2[f"B{rr}"]; bc.value=f'=IF(COUNT(חישוב!B{s}:B{e})=0,"",ROUND(AVERAGE(חישוב!B{s}:B{e}),2))'
    bc.font=Font(name=FONT,size=12,bold=True); bc.alignment=Alignment(horizontal="center",vertical="center"); bc.fill=AVG
    cc=ws2[f"C{rr}"]; cc.value=interp(f"B{rr}",rr)
    cc.font=Font(name=FONT,size=11,bold=True); cc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for col in "ABC": ws2[f"{col}{rr}"].border=border
    ws2.row_dimensions[rr].height=18; rr+=1
last_cat_row=rr-1
w(ws2,f"A{rr}","ציון כולל",13,True,"FFFFFF",fill=TOT)
tb=ws2[f"B{rr}"]; tb.value=f'=IF(COUNT(חישוב!B2:B{calc_last})=0,"",ROUND(AVERAGE(חישוב!B2:B{calc_last}),2))'
tb.font=Font(name=FONT,size=14,bold=True,color="FFFFFF"); tb.fill=TOT; tb.alignment=Alignment(horizontal="center",vertical="center")
tc=ws2[f"C{rr}"]; tc.value=interp(f"B{rr}",rr)
tc.font=Font(name=FONT,size=12,bold=True,color="FFFFFF"); tc.fill=TOT; tc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws2.row_dimensions[rr].height=24

# charts
data=Reference(ws2,min_col=2,min_row=2,max_row=last_cat_row)
catref=Reference(ws2,min_col=1,min_row=first_cat_row,max_row=last_cat_row)
bar=BarChart(); bar.type="bar"; bar.title="פרופיל מיינדסט לפי קטגוריה (1=עני · 5=עשיר)"
bar.add_data(data,titles_from_data=True); bar.set_categories(catref)
bar.y_axis.scaling.min=0; bar.y_axis.scaling.max=5; bar.x_axis.delete=False; bar.y_axis.delete=False
bar.height=12; bar.width=24; bar.legend=None; bar.dLbls=DataLabelList(); bar.dLbls.showVal=True
ws2.add_chart(bar,"E2")
radar=RadarChart(); radar.type="filled"; radar.title='מכ"ם מיינדסט'
radar.add_data(data,titles_from_data=True); radar.set_categories(catref)
radar.y_axis.scaling.min=0; radar.y_axis.scaling.max=5; radar.height=12; radar.width=14; radar.legend=None
ws2.add_chart(radar,"E26")

wb.save("/sessions/exciting-peaceful-bardeen/mnt/outputs/שאלון_מיפוי_מיינדסט.xlsx")
print("ok calc_last",calc_last)
