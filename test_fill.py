from openpyxl import load_workbook
import shutil
shutil.copy("שאלון_מיפוי_מיינדסט.xlsx","_test.xlsx")
wb=load_workbook("_test.xlsx")
ws=wb["שאלון"]
# mark: for each item row, put X in a column. Cat1 items -> poor (B=1), one rich (F=5)
marks={4:"B",5:"C",6:"D",  8:"F",  10:"E",  12:"B",13:"F",  15:"D",16:"C",17:"E"}
for r,col in marks.items():
    ws[f"{col}{r}"]="✗"
wb.save("_test.xlsx")
