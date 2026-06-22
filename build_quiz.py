# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
# colors
NAVY = "1F3864"; POOR = "C00000"; RICH = "2E5496"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
CAT_FILL = PatternFill("solid", fgColor="DCE6F1")
AVG_FILL = PatternFill("solid", fgColor="FFF2CC")
INPUT_FILL = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="1F3864")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

categories = [
 ("1. תפיסה עצמית ומקור הפרנסה", [
   "אני תופס/ת את עצמי כיוצר/ת ובעל/ת נכסים ועסק, ולא רק כאיש/אשת מקצוע שמוכר/ת את הזמן.",
   "אני שואף/ת שהכסף שלי יגיע מנכסים ומעסק, ולא רק מהמקצוע או מהמשכורת.",
   "מעניין אותי ללמוד מיומנויות עסקיות (תוכנית עסקית, גיוס משקיעים, רכישת חברות) ולא רק מקצועיות.",
   "אני לא מסתפק/ת בטיפוס בסולם הדרגות – אני רוצה לבנות משהו משלי / לשלוט בסולם.",
 ]),
 ("2. השכלה והשקעה עצמית", [
   "אני מוכן/ה להשקיע כסף וזמן בלימוד על כסף, השקעות ופיתוח אישי, ולא רק בהשכלה מקצועית.",
   "אני משקיע/ה בעצמי ובחינוך הפיננסי שלי באופן קבוע.",
 ]),
 ("3. זמן מול כסף", [
   "אני רואה בזמן שלי משאב יקר יותר מכסף, ולא ממהר/ת להחליף אותו תמורת כסף.",
   "אני מחפש/ת דרכים להרוויח כסף בלי לתת ישירות שעות עבודה תמורתו (הכנסה פאסיבית).",
 ]),
 ("4. צמיחה ולמידה", [
   "אני מאמין/ה שתמיד יש לי מה ללמוד, גם בתחומים שאני כבר מכיר/ה.",
   "כשמשהו לא מצליח, אני בודק/ת מה עליי לשנות בעצמי, במקום להניח שזה פשוט לא בשבילי.",
 ]),
 ("5. תודעת שפע מול מחסור", [
   "אני מאמין/ה שיש מספיק שפע והזדמנויות בעולם לכולם.",
   'במקום "אני לא יכול/ה להרשות לעצמי", אני שואל/ת "איך אוכל להרשות לעצמי?".',
   "כשמגיעה הזדמנות אני נוטה לפעול ולנצל אותה, ולא לחשוד ולהירתע.",
 ]),
 ("6. תפקיד הכסף – שמירה מול יצירה", [
   "אני רואה בכסף כלי שצריך לזרום ולעבוד (להשקיע ולבנות), ולא רק לשמור ולחסוך.",
   "אני מתמקד/ת ביצירת הון ותזרים, ולא רק בשמירה על מה שכבר יש לי.",
 ]),
 ("7. אסטרטגיה: הגנה מול התקפה", [
   'אני מוכן/ה לקחת סיכונים מחושבים ולהשקיע בצורה ממוקדת (אחרי לימוד), ולא רק "לשחק בטוח".',
   "אני פועל/ת מתוך ידע ולמידה עצמאית, ולא רק לפי מה שכולם עושים.",
   "אני מתעניין/ת במגוון כלים פיננסיים (מינוף, שותפויות, תזרים) ולא רק במשכורת ובחיסכון.",
 ]),
 ("8. אחריות ושליטה", [
   "אני מאמין/ה שההצלחה הכלכלית שלי תלויה בעיקר בי.",
   "כשמשהו משתבש, אני לוקח/ת אחריות ומחפש/ת מה אפשר לעשות אחרת, במקום תירוצים.",
 ]),
 ("9. תגמול לפי ערך", [
   "הייתי שמח/ה להיות מתוגמל/ת לפי הערך והתוצאות שאני מביא/ה, ולא רק לפי שעות וזמן.",
 ]),
 ("10. ערך עצמי וביטחון", [
   "אני בטוח/ה ביכולות שלי ובתוצאות שאני מסוגל/ת להשיג.",
   "הביטחון שלי נשען על היכולות שלי, ולא רק על המקצוע או התואר שלי.",
 ]),
 ("11. מיקוד ביעד מול פחד מכישלון", [
   "אני מתמקד/ת ביעד ובהזדמנות, ולא נעצר/ת בגלל פחד מכישלון.",
   "אני יוצא/ת לדרך גם כשלא כל התנאים מושלמים, וסומך/ת על עצמי לפתור בעיות בדרך.",
 ]),
 ("12. חשיבה בגדול (כמויות)", [
   'אני חושב/ת בגדול ומחפש/ת דרכים להגדיל את ההשפעה והכמות, ולא רק לעבוד "אחד על אחד".',
 ]),
 ("13. סביבה ורגשות כלפי עשירים", [
   "אני מרגיש/ה הערצה והשראה כלפי אנשים עשירים ומצליחים, ולא כעס או קנאה.",
   "אני שואף/ת להקיף את עצמי באנשים שכבר הגיעו למקום שאני רוצה להגיע אליו.",
 ]),
]

wb = Workbook()

# ---------- Sheet 1: Instructions ----------
ws0 = wb.active
ws0.title = "הוראות"
ws0.sheet_view.rightToLeft = True
ws0.column_dimensions['A'].width = 3
ws0.column_dimensions['B'].width = 100
def w(ws, cell, val, size=11, bold=False, color="000000", wrap=True, fill=None, align="right"):
    c = ws[cell]; c.value = val
    c.font = Font(name=FONT, size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill: c.fill = fill
    return c
w(ws0,"B2","שאלון מיפוי עצמי: מיינדסט עני מול מיינדסט עשיר",18,True,NAVY)
w(ws0,"B3",'"The other side of the coin"',12,False,RICH)
ws0.row_dimensions[4].height = 8
w(ws0,"B5","איך ממלאים?",13,True,NAVY)
w(ws0,"B6","עברו ללשונית \"שאלון\" ולכל היגד דרגו עד כמה הוא נכון לגביכם, בסולם מ-1 עד 5:")
w(ws0,"B7","1 = זה ממש לא אני (מיינדסט עני)   ·   3 = באמצע   ·   5 = זה בדיוק אני (מיינדסט עשיר)",11,True)
w(ws0,"B8","יש להיות כן/ה עם עצמכם – אין כאן תשובות \"נכונות\", זו מפה אישית בלבד.")
ws0.row_dimensions[9].height = 8
w(ws0,"B10","איך קוראים את התוצאה?",13,True,NAVY)
w(ws0,"B11","בתחתית לשונית \"שאלון\" (וגם בלשונית \"תוצאות\") יחושבו אוטומטית ממוצע לכל קטגוריה וציון כולל:")
w(ws0,"B12","1.0–2.0  →  מיינדסט עני דומיננטי בתחום הזה")
w(ws0,"B13","2.1–3.0  →  נטייה למיינדסט עני")
w(ws0,"B14","3.1–4.0  →  נטייה למיינדסט עשיר")
w(ws0,"B15","4.1–5.0  →  מיינדסט עשיר דומיננטי")
ws0.row_dimensions[16].height = 8
w(ws0,"B17","טיפ: שימו לב במיוחד לקטגוריות עם הציון הנמוך ביותר – שם נמצא הפוטנציאל הגדול לצמיחה.",11,True,POOR)

# ---------- Sheet 2: Questionnaire ----------
ws = wb.create_sheet("שאלון")
ws.sheet_view.rightToLeft = True
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 95
ws.column_dimensions['C'].width = 16

w(ws,"A1","שאלון מיפוי עצמי – מיינדסט עני מול מיינדסט עשיר",16,True,"FFFFFF",fill=HDR_FILL,align="center")
ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 28
for col in ("A","B","C"): ws[f"{col}1"].fill = HDR_FILL
w(ws,"A2","דרגו כל היגד: 1 = ממש לא אני (עני)  ·  5 = בדיוק אני (עשיר)",11,True,NAVY,align="center")
ws.merge_cells("A2:C2"); ws.row_dimensions[2].height = 20

# header row
hdr = 3
w(ws,f"A{hdr}","#",11,True,"FFFFFF",fill=HDR_FILL,align="center")
w(ws,f"B{hdr}","היגד",11,True,"FFFFFF",fill=HDR_FILL)
w(ws,f"C{hdr}","הדירוג שלי (1-5)",11,True,"FFFFFF",fill=HDR_FILL,align="center")
for col in ("A","B","C"): ws[f"{col}{hdr}"].border = border

dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True,
                    showErrorMessage=True, errorTitle="ערך לא חוקי", error="יש להזין מספר שלם בין 1 ל-5")
dv.prompt = "דרגו בין 1 (ממש לא אני) ל-5 (בדיוק אני)"; dv.promptTitle="דירוג"
ws.add_data_validation(dv)

r = hdr + 1
cat_avg_rows = []
all_input_rows = []
idx = 1
for cat_name, items in categories:
    w(ws,f"A{r}",cat_name,12,True,NAVY,fill=CAT_FILL)
    ws.merge_cells(f"A{r}:C{r}")
    for col in ("A","B","C"):
        ws[f"{col}{r}"].fill = CAT_FILL; ws[f"{col}{r}"].border = border
    ws.row_dimensions[r].height = 22
    r += 1
    start = r
    for it in items:
        w(ws,f"A{r}",idx,11,False,"000000",wrap=False,align="center")
        w(ws,f"B{r}",it,11)
        c = ws[f"C{r}"]; c.fill = INPUT_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name=FONT, size=12, bold=True, color="0000FF")
        for col in ("A","B","C"): ws[f"{col}{r}"].border = border
        dv.add(c)
        all_input_rows.append(r)
        ws.row_dimensions[r].height = 30
        idx += 1; r += 1
    end = r - 1
    # category average row
    w(ws,f"B{r}","ממוצע הקטגוריה",11,True,"7F6000",align="right",fill=AVG_FILL)
    cf = ws[f"C{r}"]
    cf.value = f'=IF(COUNT(C{start}:C{end})=0,"",ROUND(AVERAGE(C{start}:C{end}),2))'
    cf.font = Font(name=FONT, size=12, bold=True, color="7F6000")
    cf.alignment = Alignment(horizontal="center", vertical="center")
    cf.fill = AVG_FILL
    ws[f"A{r}"].fill = AVG_FILL
    for col in ("A","B","C"): ws[f"{col}{r}"].border = border
    cat_avg_rows.append((cat_name, r))
    ws.row_dimensions[r].height = 20
    r += 1

# total score
r += 1
w(ws,f"B{r}","ציון כולל (ממוצע כל ההיגדים)",13,True,"FFFFFF",fill=TOTAL_FILL,align="right")
first_in = all_input_rows[0]; last_in = all_input_rows[-1]
tot = ws[f"C{r}"]
tot.value = f'=IF(COUNT(C{first_in}:C{last_in})=0,"",ROUND(AVERAGE(C{first_in}:C{last_in}),2))'
tot.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
tot.alignment = Alignment(horizontal="center", vertical="center")
tot.fill = TOTAL_FILL
ws[f"A{r}"].fill = TOTAL_FILL
for col in ("A","B","C"): ws[f"{col}{r}"].border = border
ws.row_dimensions[r].height = 26
total_row = r
r += 1
w(ws,f"B{r}","פירוש הציון",12,True,NAVY,align="right",fill=CAT_FILL)
pc = ws[f"C{r}"]
pc.value = (f'=IF(C{total_row}="","-",'
            f'IF(C{total_row}<=2,"מיינדסט עני דומיננטי",'
            f'IF(C{total_row}<=3,"נטייה למיינדסט עני",'
            f'IF(C{total_row}<=4,"נטייה למיינדסט עשיר","מיינדסט עשיר דומיננטי"))))')
pc.font = Font(name=FONT, size=12, bold=True, color=POOR)
pc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
pc.fill = CAT_FILL
ws[f"A{r}"].fill = CAT_FILL
for col in ("A","B","C"): ws[f"{col}{r}"].border = border
ws.freeze_panes = "A4"

# ---------- Sheet 3: Results summary ----------
ws2 = wb.create_sheet("תוצאות")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 28
w(ws2,"A1","סיכום תוצאות לפי קטגוריה",16,True,"FFFFFF",fill=HDR_FILL,align="center")
ws2.merge_cells("A1:C1"); ws2.row_dimensions[1].height = 26
for col in ("A","B","C"): ws2[f"{col}1"].fill = HDR_FILL
w(ws2,"A2","קטגוריה",11,True,"FFFFFF",fill=HDR_FILL)
w(ws2,"B2","ממוצע",11,True,"FFFFFF",fill=HDR_FILL,align="center")
w(ws2,"C2","פירוש",11,True,"FFFFFF",fill=HDR_FILL,align="center")
for col in ("A","B","C"): ws2[f"{col}2"].border = border
rr = 3
for cat_name, avg_row in cat_avg_rows:
    w(ws2,f"A{rr}",cat_name,11)
    bc = ws2[f"B{rr}"]; bc.value = f"=שאלון!C{avg_row}"
    bc.font = Font(name=FONT, size=12, bold=True, color="000000")
    bc.alignment = Alignment(horizontal="center", vertical="center")
    cc = ws2[f"C{rr}"]
    cc.value = (f'=IF(B{rr}="","-",'
                f'IF(B{rr}<=2,"מיינדסט עני דומיננטי",'
                f'IF(B{rr}<=3,"נטייה למיינדסט עני",'
                f'IF(B{rr}<=4,"נטייה למיינדסט עשיר","מיינדסט עשיר דומיננטי"))))')
    cc.font = Font(name=FONT, size=11, bold=True)
    cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ("A","B","C"): ws2[f"{col}{rr}"].border = border
    ws2.row_dimensions[rr].height = 20
    rr += 1
# total
w(ws2,f"A{rr}","ציון כולל",13,True,"FFFFFF",fill=TOTAL_FILL)
tb = ws2[f"B{rr}"]; tb.value = f"=שאלון!C{total_row}"
tb.font = Font(name=FONT, size=14, bold=True, color="FFFFFF"); tb.fill = TOTAL_FILL
tb.alignment = Alignment(horizontal="center", vertical="center")
tc = ws2[f"C{rr}"]
tc.value = (f'=IF(B{rr}="","-",'
            f'IF(B{rr}<=2,"מיינדסט עני דומיננטי",'
            f'IF(B{rr}<=3,"נטייה למיינדסט עני",'
            f'IF(B{rr}<=4,"נטייה למיינדסט עשיר","מיינדסט עשיר דומיננטי"))))')
tc.font = Font(name=FONT, size=12, bold=True, color="FFFFFF"); tc.fill = TOTAL_FILL
tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[rr].height = 26

wb.save("/sessions/exciting-peaceful-bardeen/mnt/outputs/שאלון_מיפוי_מיינדסט.xlsx")
print("saved", total_row)
