# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT="Arial"; NAVY="1F3864"; POOR="C00000"; RICH="2E5496"
HDR=PatternFill("solid",fgColor="1F3864")
POORH=PatternFill("solid",fgColor="F4CCCC")
RICHH=PatternFill("solid",fgColor="D9E1F2")
CAT=PatternFill("solid",fgColor="EDEDED")
AVG=PatternFill("solid",fgColor="FFF2CC")
TOT=PatternFill("solid",fgColor="1F3864")
INP=PatternFill("solid",fgColor="FFFFFF")
thin=Side(style="thin",color="BFBFBF"); border=Border(thin,thin,thin,thin)

# (category, [(poor, rich), ...])
cats_data=[
 ("1. תפיסה עצמית ומקור הפרנסה",[
   ('"אני איש/אשת מקצוע" – הכסף שלי מגיע מהמקצוע שלי.','"אני בעל/ת עסק" – הכסף שלי מגיע מהעסק ומהנכסים שלי.'),
   ("חשוב לי ללמוד איך כותבים קורות חיים ומתכוננים לראיון עבודה.","חשוב לי ללמוד איך כותבים תוכנית עסקית, מגייסים משקיעים ורוכשים חברות."),
   ("השאיפה שלי היא להתקדם בסולם הדרגות.","השאיפה שלי היא לשלוט בסולם / לבנות משהו משלי."),
 ]),
 ("2. השכלה והשקעה עצמית",[
   ("משקיע/ה בהשכלה המקצועית שלי, אבל פחות בעצמי ובידע על כסף.","משקיע/ה בעצמי, בחינוך הפיננסי ובמיומנויות של כסף וניהול."),
 ]),
 ("3. זמן מול כסף",[
   ('"הזמן שלי שווה כסף" – אני מחליף/ה זמן בכסף (מוכר/ת שעות).','"הזמן שלי יקר מכסף" – אני מחליף/ה זמן ביצירת נכסים והכנסה פאסיבית.'),
 ]),
 ("4. צמיחה ולמידה",[
   ("בתחומים שאני מכיר/ה אני כבר יודע/ת מספיק, אין הרבה מה ללמוד.","תמיד יש לי מה ללמוד – \"His cup is never full\"."),
   ("כשמשהו לא מצליח, כנראה שזה פשוט לא בשבילי.","כשמשהו לא מצליח, אני בודק/ת מה לשנות בעצמי."),
 ]),
 ("5. תודעת שפע מול מחסור",[
   ("כמות הכסף בעולם מוגבלת – אם אחד מרוויח, אחר מפסיד.","יש שפע והזדמנויות לכולם – הכסף גדל ככל שיש יותר שחקנים."),
   ('"אני לא יכול/ה להרשות לעצמי את זה".','"איך אוכל להרשות לעצמי את זה?".'),
   ("כשמגיעה הזדמנות אני חשדן/ית ונרתע/ת.","כשמגיעה הזדמנות אני פועל/ת ומנצל/ת אותה."),
 ]),
 ("6. תפקיד הכסף – שמירה מול יצירה",[
   ("כסף נועד בעיקר להישמר ולהיחסך.","כסף נועד לזרום, להיות מושקע ולבנות."),
 ]),
 ("7. אסטרטגיה: הגנה מול התקפה",[
   ('משחק/ת "הגנה" – לא רוצה להפסיד, נמנע/ת מסיכונים, משחק/ת בטוח.','משחק/ת "התקפה" – רוצה לנצח, לוקח/ת סיכונים מחושבים.'),
   ("פועל/ת לפי מה שכולם עושים ומתייעץ/ת עם הסביבה.","פועל/ת מתוך ידע ולמידה עצמאית (בד\"כ הפוך מהרוב)."),
 ]),
 ("8. אחריות ושליטה",[
   ("ההצלחה הכלכלית שלי תלויה בגורמים שאין לי שליטה עליהם.","ההצלחה הכלכלית שלי תלויה בעיקר בי."),
   ("כשמשהו משתבש אני מחפש/ת את הסיבה החיצונית / האשם.","כשמשהו משתבש אני לוקח/ת אחריות ומחפש/ת מה לשנות."),
 ]),
 ("9. תגמול לפי ערך",[
   ("מעדיף/ה שכר קבוע לפי זמן ושעות.","מעדיף/ה תגמול לפי הערך והתוצאות שאני מביא/ה."),
 ]),
 ("10. ערך עצמי וביטחון",[
   ("לא תמיד בטוח/ה ביכולות שלי ובתוצאות שאוכל להשיג.","בטוח/ה ביכולות שלי ובתוצאות שאני מסוגל/ת להשיג."),
   ("הביטחון שלי נשען על המקצוע / התואר שלי.","הביטחון שלי נשען על היכולות שלי."),
 ]),
 ("11. מיקוד ביעד מול פחד מכישלון",[
   ("פחד מכישלון עוצר אותי – אפעל רק כשהכול מושלם.","מתמקד/ת ביעד, יוצא/ת לדרך גם כשלא הכול מושלם."),
   ("ניתוח יתר נוטה לשתק אותי (analysis paralysis).","הניתוח עוזר לי לבחור, להבין סיכונים ולפעול."),
 ]),
 ("12. חשיבה בגדול (כמויות)",[
   ('חושב/ת ועובד/ת בקטן, בעיקר "אחד על אחד".',"חושב/ת בגדול, מחפש/ת להגדיל את הכמות וההשפעה."),
 ]),
 ("13. סביבה ורגשות כלפי עשירים",[
   ("כלפי אנשים עשירים אני נוטה להרגיש כעס / קנאה / ביקורת.","כלפי אנשים עשירים אני מרגיש/ה הערצה והשראה."),
   ("נוח לי בסביבה של אנשים עם מיינדסט דומה לשלי.","שואף/ת להקיף את עצמי באנשים שהגיעו לאן שאני רוצה."),
 ]),
]

wb=Workbook()
def w(ws,cell,val,size=11,bold=False,color="000000",wrap=True,fill=None,align="right",valign="center"):
    c=ws[cell]; c.value=val
    c.font=Font(name=FONT,size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal=align,vertical=valign,wrap_text=wrap)
    if fill:c.fill=fill
    return c

# ---- הוראות ----
ws0=wb.active; ws0.title="הוראות"; ws0.sheet_view.rightToLeft=True
ws0.column_dimensions['A'].width=3; ws0.column_dimensions['B'].width=105
w(ws0,"B2","שאלון מיפוי עצמי: מיינדסט עני מול מיינדסט עשיר",18,True,NAVY)
w(ws0,"B3",'"The other side of the coin"',12,False,RICH)
w(ws0,"B5","איך ממלאים?",13,True,NAVY)
w(ws0,"B6","בלשונית \"שאלון\", כל שורה מציגה שני ניסוחים מנוגדים: מימין ניסוח של מיינדסט עני, משמאל ניסוח של מיינדסט עשיר.")
w(ws0,"B7","עבור כל שורה, סמנו בעמודה האמצעית מספר מ-1 עד 5 לפי מה שמתאר אתכם באמת:")
w(ws0,"B8","1 = מתחבר/ת לחלוטין לניסוח הימני (עני)   ·   2 = נוטה לימני   ·   3 = שילוב / באמצע   ·   4 = נוטה לשמאלי   ·   5 = מתחבר/ת לחלוטין לניסוח השמאלי (עשיר)",11,True)
w(ws0,"B9","אין כאן תשובה \"נכונה\" – שני הצדדים מוצגים באופן שווה. ענו בכנות, זו מפה אישית בלבד.")
w(ws0,"B11","איך קוראים את התוצאה?",13,True,NAVY)
w(ws0,"B12","בתחתית \"שאלון\" וגם בלשונית \"תוצאות\" יחושבו אוטומטית ממוצע לכל קטגוריה וציון כולל:")
w(ws0,"B13","1.0–2.0  →  מיינדסט עני דומיננטי")
w(ws0,"B14","2.1–3.0  →  נטייה למיינדסט עני")
w(ws0,"B15","3.1–4.0  →  נטייה למיינדסט עשיר")
w(ws0,"B16","4.1–5.0  →  מיינדסט עשיר דומיננטי")
w(ws0,"B18","טיפ: התמקדו בקטגוריות עם הציון הנמוך ביותר – שם הפוטנציאל הגדול ביותר לצמיחה.",11,True,POOR)

# ---- שאלון ----
ws=wb.create_sheet("שאלון"); ws.sheet_view.rightToLeft=True
ws.column_dimensions['A'].width=58; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=58
w(ws,"A1","שאלון מיפוי עצמי – מיינדסט עני מול מיינדסט עשיר",15,True,"FFFFFF",fill=HDR,align="center")
ws.merge_cells("A1:C1"); ws.row_dimensions[1].height=26
for col in "ABC": ws[f"{col}1"].fill=HDR
w(ws,"A2","בכל שורה: דרגו 1-5.  1 = מתחבר/ת לימני (עני)  ·  3 = שילוב  ·  5 = מתחבר/ת לשמאלי (עשיר)",11,True,NAVY,align="center")
ws.merge_cells("A2:C2"); ws.row_dimensions[2].height=20
hdr=3
w(ws,f"A{hdr}","מיינדסט עני  (1)",12,True,POOR,fill=POORH,align="center")
w(ws,f"B{hdr}","הדירוג שלי",11,True,"FFFFFF",fill=HDR,align="center")
w(ws,f"C{hdr}","מיינדסט עשיר  (5)",12,True,RICH,fill=RICHH,align="center")
for col in "ABC": ws[f"{col}{hdr}"].border=border
ws.row_dimensions[hdr].height=20

dv=DataValidation(type="whole",operator="between",formula1="1",formula2="5",allow_blank=True,
                  showErrorMessage=True,errorTitle="ערך לא חוקי",error="יש להזין מספר שלם בין 1 ל-5")
dv.prompt="1=עני … 5=עשיר"; dv.promptTitle="דירוג"; ws.add_data_validation(dv)

r=hdr+1; cat_avg_rows=[]; inp_rows=[]
for cat_name,pairs in cats_data:
    w(ws,f"A{r}",cat_name,12,True,NAVY,fill=CAT,align="center")
    ws.merge_cells(f"A{r}:C{r}")
    for col in "ABC": ws[f"{col}{r}"].fill=CAT; ws[f"{col}{r}"].border=border
    ws.row_dimensions[r].height=20; r+=1
    start=r
    for poor,rich in pairs:
        w(ws,f"A{r}",poor,11,color="000000")
        c=ws[f"B{r}"]; c.fill=INP; c.border=border
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.font=Font(name=FONT,size=13,bold=True,color="0000FF"); dv.add(c)
        w(ws,f"C{r}",rich,11,color="000000")
        ws[f"A{r}"].border=border; ws[f"C{r}"].border=border
        ws.row_dimensions[r].height=42; inp_rows.append(r); r+=1
    end=r-1
    w(ws,f"A{r}","ממוצע הקטגוריה  ←",11,True,"7F6000",fill=AVG,align="left")
    cf=ws[f"B{r}"]; cf.value=f'=IF(COUNT(B{start}:B{end})=0,"",ROUND(AVERAGE(B{start}:B{end}),2))'
    cf.font=Font(name=FONT,size=12,bold=True,color="7F6000")
    cf.alignment=Alignment(horizontal="center",vertical="center"); cf.fill=AVG
    w(ws,f"C{r}","",fill=AVG)
    for col in "ABC": ws[f"{col}{r}"].border=border
    cat_avg_rows.append((cat_name,r)); ws.row_dimensions[r].height=18; r+=1

r+=1
w(ws,f"A{r}","ציון כולל (ממוצע כל ההיגדים)  ←",13,True,"FFFFFF",fill=TOT,align="left")
tot=ws[f"B{r}"]
tot.value=f'=IF(COUNT(B{inp_rows[0]}:B{inp_rows[-1]})=0,"",ROUND(AVERAGE(B{inp_rows[0]}:B{inp_rows[-1]}),2))'
tot.font=Font(name=FONT,size=14,bold=True,color="FFFFFF"); tot.fill=TOT
tot.alignment=Alignment(horizontal="center",vertical="center")
w(ws,f"C{r}","",fill=TOT)
for col in "ABC": ws[f"{col}{r}"].border=border
ws.row_dimensions[r].height=24; total_row=r; r+=1
w(ws,f"A{r}","פירוש הציון  ←",12,True,NAVY,fill=CAT,align="left")
pc=ws[f"B{r}"]
pc.value=(f'=IF(B{total_row}="","-",IF(B{total_row}<=2,"עני דומיננטי",'
          f'IF(B{total_row}<=3,"נטייה לעני",IF(B{total_row}<=4,"נטייה לעשיר","עשיר דומיננטי"))))')
pc.font=Font(name=FONT,size=11,bold=True,color=POOR)
pc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); pc.fill=CAT
w(ws,f"C{r}","",fill=CAT)
for col in "ABC": ws[f"{col}{r}"].border=border
ws.freeze_panes="A4"

# ---- תוצאות ----
ws2=wb.create_sheet("תוצאות"); ws2.sheet_view.rightToLeft=True
ws2.column_dimensions['A'].width=42; ws2.column_dimensions['B'].width=12; ws2.column_dimensions['C'].width=24
w(ws2,"A1","סיכום תוצאות לפי קטגוריה",15,True,"FFFFFF",fill=HDR,align="center")
ws2.merge_cells("A1:C1"); ws2.row_dimensions[1].height=24
for col in "ABC": ws2[f"{col}1"].fill=HDR
w(ws2,"A2","קטגוריה",11,True,"FFFFFF",fill=HDR)
w(ws2,"B2","ממוצע",11,True,"FFFFFF",fill=HDR,align="center")
w(ws2,"C2","פירוש",11,True,"FFFFFF",fill=HDR,align="center")
for col in "ABC": ws2[f"{col}2"].border=border
rr=3
for cat_name,avg_row in cat_avg_rows:
    w(ws2,f"A{rr}",cat_name,11)
    bc=ws2[f"B{rr}"]; bc.value=f"=שאלון!B{avg_row}"
    bc.font=Font(name=FONT,size=12,bold=True); bc.alignment=Alignment(horizontal="center",vertical="center")
    cc=ws2[f"C{rr}"]
    cc.value=(f'=IF(B{rr}="","-",IF(B{rr}<=2,"עני דומיננטי",IF(B{rr}<=3,"נטייה לעני",'
              f'IF(B{rr}<=4,"נטייה לעשיר","עשיר דומיננטי"))))')
    cc.font=Font(name=FONT,size=11,bold=True); cc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for col in "ABC": ws2[f"{col}{rr}"].border=border
    ws2.row_dimensions[rr].height=18; rr+=1
w(ws2,f"A{rr}","ציון כולל",13,True,"FFFFFF",fill=TOT)
tb=ws2[f"B{rr}"]; tb.value=f"=שאלון!B{total_row}"
tb.font=Font(name=FONT,size=14,bold=True,color="FFFFFF"); tb.fill=TOT
tb.alignment=Alignment(horizontal="center",vertical="center")
tc=ws2[f"C{rr}"]
tc.value=(f'=IF(B{rr}="","-",IF(B{rr}<=2,"עני דומיננטי",IF(B{rr}<=3,"נטייה לעני",'
          f'IF(B{rr}<=4,"נטייה לעשיר","עשיר דומיננטי"))))')
tc.font=Font(name=FONT,size=12,bold=True,color="FFFFFF"); tc.fill=TOT
tc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws2.row_dimensions[rr].height=24
import json; print(json.dumps({"total_row":total_row,"last_cat":rr-1}))
wb.save("/sessions/exciting-peaceful-bardeen/mnt/outputs/שאלון_מיפוי_מיינדסט.xlsx")
